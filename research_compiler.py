import csv
from APIs.BrowseAI import *
from HelperFunctions import *
import time
import openpyxl

BrowseAI_API_KEY = get_secret_api_key()
BrowseAI_Google_Search_Robot_ID = get_robot_app("Google Search")
BrowseAI_Google_Scholars_Robot_ID = get_robot_app("Google Scholar")

def main():
    print("(--------------------- Welcome to Research Compiler ---------------------)")
    print("")
    print("* A program that can help you find the answer to your research questions *")
    print("")
    print("Give me a search term and I'll give you a list of things you might want to know about it.")
    print("")
    search_term = input("What do you want to dig up on?: ")
    print("")
    print("Alright! I'm searching for " + search_term + "...")
    print("This might take a while...")
    send_info(search_term)
    print("Waiting for BrowseAI to finish... (est 90 seconds)")
    for i in range(90, 0, -1):
        print("Finished in " + str(i) + " seconds...")
        time.sleep(1)
    results = get_info(search_term)
    print("Okay! I found some information about " + search_term + "!")
    print("Exporting it to a Excel file...")
    write_to_csv(search_term, results)
    print("Done! I exported it to a file called " + search_term + ".xlsx...")
    file_location = r"C:\Users\Nabil\Desktop\ResearchCompiler" + "\\" + search_term + ".xlsx"
    print("All set! The file location is at " + file_location + ".")
    exit = input("Press any key to exit...")

def send_info(search_term):
    global BrowseAI_Google_Search_Robot_ID, BrowseAI_Google_Scholars_Robot_ID, BrowseAI_API_KEY
    # BrowseAI #
    create_robot_task(BrowseAI_Google_Search_Robot_ID, {"inputParameters": {"originUrl" : search_term, "Organic Result_limit": 20}}, BrowseAI_API_KEY)
    time.sleep(2)
    create_robot_task(BrowseAI_Google_Scholars_Robot_ID, {"inputParameters": {"search_keyword" : search_term, "articles_list_limit": 20}}, BrowseAI_API_KEY)
    time.sleep(2)
    print("Sending requests to BROWSE AI...")

def get_info(search_term):
    print("(--------------------- Retrieving basic information ---------------------)")
    print("Getting basic information about " + search_term + "...")
    definition = "https://www.dictionary.com/browse/" + search_term
    wikipedia =  "https://en.wikipedia.org/wiki/" + search_term
    Google_News = search_articles(search_term,2)
    time.sleep(0.2)
    print("Retrieving results from BROWSE AI...")
    GoogleSearchResults = get_latest_robot_task(BrowseAI_Google_Search_Robot_ID,BrowseAI_API_KEY)
    GoogleScholarsResults = get_latest_robot_task(BrowseAI_Google_Scholars_Robot_ID,BrowseAI_API_KEY)
    print("Processing results from BROWSE AI...")
    GSE_Results = process_robot_task_data(GoogleSearchResults)
    GSC_Results = process_robot_task_data(GoogleScholarsResults)
    return definition, wikipedia, GSE_Results, GSC_Results, Google_News


def write_to_csv(file_name, data):
    # Create a new workbook and select the active sheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Results"

    # Define the values to be added to the first two rows
    search = "Search Term:"
    value = file_name

    value1 = "Definition:"
    value2 = data[0]
    value3 = "Wikipedia:"
    value4 = data[1]

    # Write the values to the first two rows
    sheet.cell(row=1, column=1, value=search)
    sheet.cell(row=1, column=2, value=value)
    sheet.cell(row=2, column=1, value=value1)
    sheet.cell(row=2, column=2, value=value2)
    sheet.cell(row=3, column=1, value=value3)
    sheet.cell(row=3, column=2, value=value4)

    # Define the nested lists
    GSE = data[2]
    GSC = data[3]
    GN = data[4]
    GSE_titles = [['Position','Title','Description','Link']]
    GSC_titles = [['Position','Title','Author','Description','Article','PDF']]
    sheet.cell(row=1, column=4, value="Google Search Results")
    for row_index, row_values in enumerate(GSE_titles, start=2):
        for col_index, col_value in enumerate(row_values, start=4):
            sheet.cell(row=row_index, column=col_index, value=col_value)
    
    # Iterate over each sublist in the nested list and write the values to the Excel file
    for row_index, row_values in enumerate(GSE, start=3):
        for col_index, col_value in enumerate(row_values, start=4):
            sheet.cell(row=row_index, column=col_index, value=col_value)

    sheet.cell(row=23, column=4, value="Google Scholars Results")
    for row_index, row_values in enumerate(GSC_titles, start=24):
        for col_index, col_value in enumerate(row_values, start=4):
            sheet.cell(row=row_index, column=col_index, value=col_value)

    for row_index, row_values in enumerate(GSC, start=25):
        for col_index, col_value in enumerate(row_values, start=4):
            sheet.cell(row=row_index, column=col_index, value=col_value)


    sheet.cell(row=45, column=4, value="Google News Results")
    sheet.cell(row=46, column=4, value="No")
    sheet.cell(row=46, column=5, value="Links")
    start_row = 47
    start_col = 4
    for i, link in enumerate(GN):
        row = start_row + i
        sheet.cell(row=row, column=start_col, value=str(i))
        col = start_col + 1
        sheet.cell(row=row, column=col, value=link)

    # Autoformat the rows and columns
    for row in sheet.rows:
        max_height = 0
        for cell in row:
            try:
                if len(str(cell.value)) > max_height:
                    max_height = len(str(cell.value))
            except:
                pass
        sheet.row_dimensions[cell.row].height = 15
    for column in sheet.columns:
        max_length = 0
        column_name = openpyxl.utils.cell.column_index_from_string(column[0].column_letter)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        sheet.column_dimensions[openpyxl.utils.cell.get_column_letter(column_name)].width = adjusted_width


    # Save the workbook to a file
    workbook.save(file_name + '.xlsx')
main()